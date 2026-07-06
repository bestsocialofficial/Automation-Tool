"""
dashboard.py

FastAPI backend + static frontend for the keyword ranking dashboard.
Reads the JSON file produced by semrush_scraper.py and serves:

    GET /                    the dashboard UI
    GET /api/dashboard       all computed dashboard data (KPIs, trend,
                             keyword table, distributions, winners/losers)

Run:
    python dashboard.py                                  # uses results.json
    python dashboard.py --data demo_results.json         # demo dataset
    python dashboard.py --port 9000
"""

import argparse
import io
import json
import os
import secrets
from collections import defaultdict
from datetime import datetime
from statistics import mean
from typing import Optional

from dotenv import load_dotenv
from fastapi import Depends, FastAPI, HTTPException, Request
from fastapi.responses import FileResponse, RedirectResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from starlette.middleware.sessions import SessionMiddleware

from auth_utils import verify_password

load_dotenv()

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STATIC_DIR = os.path.join(BASE_DIR, "static")

# Default source is MongoDB. Set --data (or the RESULTS_FILE env var) to a
# JSON file to serve from a file instead — used for the demo dataset.
DATA_FILE = os.environ.get("RESULTS_FILE") or None

app = FastAPI(title="Keyword Ranking Dashboard")

_mongo_client = None


def get_db():
    global _mongo_client
    from pymongo import MongoClient

    uri, db_name, _ = mongo_settings()
    if _mongo_client is None:
        _mongo_client = MongoClient(uri, serverSelectionTimeoutMS=5000)
    return _mongo_client[db_name]


def _session_secret():
    """Signing key for session cookies — persisted in Mongo so logins
    survive server restarts. SECRET_KEY in .env overrides it."""
    env = os.environ.get("SECRET_KEY")
    if env:
        return env
    try:
        config = get_db()["config"]
        doc = config.find_one({"_id": "session_secret"})
        if doc is None:
            value = secrets.token_hex(32)
            config.insert_one({"_id": "session_secret", "value": value})
            return value
        return doc["value"]
    except Exception:
        # Mongo unreachable right now — fall back to a per-run key
        # (existing logins won't survive a restart).
        return secrets.token_hex(32)


app.add_middleware(
    SessionMiddleware,
    secret_key=_session_secret(),
    max_age=7 * 24 * 3600,  # stay signed in for a week
    same_site="lax",
)

BUCKET_LABELS = ["Top 3", "4-10", "11-20", "21-50", "51-100", "Not ranking"]


def mongo_settings():
    return (
        os.environ.get("MONGO_URI", "mongodb://localhost:27017"),
        os.environ.get("MONGO_DB", "seo"),
        os.environ.get("MONGO_COLLECTION", "keyword_rankings"),
    )


def load_records():
    if DATA_FILE:
        if not os.path.exists(DATA_FILE):
            return []
        with open(DATA_FILE, "r", encoding="utf-8") as fh:
            content = fh.read().strip()
        return json.loads(content) if content else []

    coll_name = mongo_settings()[2]
    docs = list(get_db()[coll_name].find({}, {"_id": 0}))
    # Mongo stores fetch_date as a datetime; the rest of the code expects
    # the ISO string format the scraper uses.
    for doc in docs:
        if isinstance(doc.get("fetch_date"), datetime):
            doc["fetch_date"] = doc["fetch_date"].strftime("%Y-%m-%dT%H:%M:%SZ")
    return docs


def date_of(record):
    return (record.get("fetch_date") or "")[:10]


def bucket_of(position):
    if position is None or position > 100:
        return "Not ranking"
    if position <= 3:
        return "Top 3"
    if position <= 10:
        return "4-10"
    if position <= 20:
        return "11-20"
    if position <= 50:
        return "21-50"
    return "51-100"


def build_series(records, domain):
    """keyword -> {date: record}. Later records in the file win for the
    same date, so a re-run on the same day overwrites the earlier fetch."""
    series = defaultdict(dict)
    for record in records:
        if record.get("domain") != domain:
            continue
        day = date_of(record)
        if day and record.get("keyword"):
            series[record["keyword"]][day] = record
    return series


def build_trend(series):
    """Per-date aggregates across all keywords fetched on that date."""
    all_dates = sorted({d for dates in series.values() for d in dates})
    trend = []
    for day in all_dates:
        positions = [
            s[day]["position"]
            for s in series.values()
            if day in s and s[day].get("position") is not None
        ]
        tracked = sum(1 for s in series.values() if day in s)
        trend.append(
            {
                "date": day,
                "tracked": tracked,
                "ranking": len(positions),
                "avg_position": round(mean(positions), 1) if positions else None,
                "top3": sum(1 for p in positions if p <= 3),
                "top10": sum(1 for p in positions if p <= 10),
            }
        )
    return trend


def build_keywords(series):
    """Latest snapshot per keyword, with change vs its previous fetch."""
    keywords = []
    for keyword, by_date in series.items():
        dates = sorted(by_date)
        latest = by_date[dates[-1]]
        prev = by_date[dates[-2]] if len(dates) > 1 else None

        position = latest.get("position")
        prev_position = prev.get("position") if prev else None
        change = None
        is_new = False
        if prev is not None:
            if prev_position is not None and position is not None:
                change = prev_position - position  # positive = improved
            elif prev_position is None and position is not None:
                is_new = True

        keywords.append(
            {
                "keyword": keyword,
                "date": dates[-1],
                "position": position,
                "prev_position": prev_position,
                "change": change,
                "is_new": is_new,
                "search_volume": latest.get("search_volume"),
                "traffic": latest.get("traffic"),
                "keyword_difficulty": latest.get("keyword_difficulty"),
                "search_intent": latest.get("search_intent"),
                "vi": latest.get("vi"),
                "bucket": bucket_of(position),
                "history": [
                    {"date": d, "position": by_date[d].get("position")} for d in dates
                ],
            }
        )
    # Best positions first, non-ranking last, then by volume.
    keywords.sort(
        key=lambda k: (
            k["position"] is None,
            k["position"] if k["position"] is not None else 0,
            -(k["search_volume"] or 0),
        )
    )
    return keywords


def build_kpis(trend, keywords):
    latest = trend[-1] if trend else None
    prev = trend[-2] if len(trend) > 1 else None

    def delta(field, invert=False):
        if not latest or not prev:
            return None
        a, b = latest.get(field), prev.get(field)
        if a is None or b is None:
            return None
        d = a - b
        return round(-d if invert else d, 1)

    return {
        "total_keywords": len(keywords),
        "ranking": sum(1 for k in keywords if k["position"] is not None),
        "top3": latest["top3"] if latest else 0,
        "top10": latest["top10"] if latest else 0,
        "avg_position": latest["avg_position"] if latest else None,
        "total_volume": sum(k["search_volume"] or 0 for k in keywords),
        "deltas": {
            "top3": delta("top3"),
            "top10": delta("top10"),
            # For positions, lower is better -> invert so positive = improved.
            "avg_position": delta("avg_position", invert=True),
        },
    }


def build_movers(keywords):
    changed = [k for k in keywords if k["change"] is not None]
    # Newly-ranking keywords lead the winners list, then biggest jumps.
    winners = [k for k in keywords if k["is_new"]]
    winners.extend(
        sorted((k for k in changed if k["change"] > 0), key=lambda k: -k["change"])
    )
    losers = sorted(
        (k for k in changed if k["change"] < 0), key=lambda k: k["change"]
    )[:5]
    slim = lambda k: {
        "keyword": k["keyword"],
        "position": k["position"],
        "change": k["change"],
        "is_new": k["is_new"],
    }
    return [slim(k) for k in winners[:5]], [slim(k) for k in losers]


# ---------------------------------------------------------------- auth


class LoginBody(BaseModel):
    username: str
    password: str


def current_user(request: Request):
    username = request.session.get("user")
    if not username:
        raise HTTPException(status_code=401, detail="Not signed in")
    user = get_db()["users"].find_one({"username": username})
    if not user:
        request.session.clear()
        raise HTTPException(status_code=401, detail="Not signed in")
    return user


@app.post("/api/login")
def api_login(body: LoginBody, request: Request):
    user = get_db()["users"].find_one({"username": body.username.strip().lower()})
    if not user or not verify_password(body.password, user["salt"], user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid username or password")
    request.session["user"] = user["username"]
    return {"ok": True}


@app.post("/api/logout")
def api_logout(request: Request):
    request.session.clear()
    return {"ok": True}


@app.get("/api/me")
def api_me(user=Depends(current_user)):
    return {
        "username": user["username"],
        "is_admin": bool(user.get("is_admin")),
        "domains": user.get("domains") or [],
    }


def allowed_domains(user, all_domains):
    """Admins and users without a domain list see everything."""
    if user.get("is_admin") or not user.get("domains"):
        return all_domains
    allowed = set(user["domains"])
    return [d for d in all_domains if d in allowed]


# ---------------------------------------------------------------- data api


@app.get("/api/dashboard")
def dashboard_data(domain: Optional[str] = None, user=Depends(current_user)):
    records = load_records()
    if not records:
        return {"empty": True, "domains": []}

    # Domains ordered by most recent fetch, most recent first.
    last_seen = {}
    for record in records:
        d = record.get("domain")
        if d:
            last_seen[d] = max(last_seen.get(d, ""), date_of(record))
    all_domains = sorted(last_seen, key=lambda d: last_seen[d], reverse=True)
    domains = allowed_domains(user, all_domains)
    if not domains:
        return {"empty": True, "domains": []}
    if domain not in domains:
        domain = domains[0]

    series = build_series(records, domain)
    trend = build_trend(series)
    keywords = build_keywords(series)
    winners, losers = build_movers(keywords)

    buckets = {label: 0 for label in BUCKET_LABELS}
    intents = defaultdict(int)
    for k in keywords:
        buckets[k["bucket"]] += 1
        intent = k["search_intent"]
        for label in intent if isinstance(intent, list) else [intent]:
            if label:
                intents[label] += 1

    latest_records = [k for k in keywords if k["date"] == trend[-1]["date"]]
    database = next(
        (
            series[k["keyword"]][k["date"]].get("database")
            for k in latest_records
            if series[k["keyword"]][k["date"]].get("database")
        ),
        None,
    )

    return {
        "empty": False,
        "domains": domains,
        "domain": domain,
        "database": database,
        "last_updated": trend[-1]["date"] if trend else None,
        "kpis": build_kpis(trend, keywords),
        "trend": trend,
        "keywords": keywords,
        "buckets": buckets,
        "intents": dict(intents),
        "winners": winners,
        "losers": losers,
    }


@app.get("/api/export.xlsx")
def export_xlsx(domain: Optional[str] = None, user=Depends(current_user)):
    data = dashboard_data(domain, user)
    if data.get("empty"):
        raise HTTPException(status_code=404, detail="No data to export")

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Keyword Report"

    ws["A1"] = f"Keyword ranking report \u2014 {data['domain']}"
    ws["A1"].font = Font(bold=True, size=14)
    market = f"Google {data['database'].upper()}" if data.get("database") else "-"
    ws["A2"] = (
        f"Market: {market}    Last updated: {data['last_updated']}    Source: SEMrush"
    )
    ws["A2"].font = Font(color="66738A")

    headers = ["Keyword", "Position", "Previous Position", "Change", "Search Volume",
               "Traffic %", "Keyword Difficulty", "Intent", "Est. Visibility", "Last Checked"]
    header_row = 4
    header_fill = PatternFill("solid", start_color="4F5DF0")
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    green, red = Font(color="14893C"), Font(color="CC3340")
    for i, k in enumerate(data["keywords"], start=header_row + 1):
        intent = k["search_intent"]
        if isinstance(intent, list):
            intent = ", ".join(intent)
        position = k["position"] if k["position"] is not None else "Not in top 100"
        change = "New" if k["is_new"] else k["change"]
        row = [k["keyword"], position, k["prev_position"], change,
               k["search_volume"], k["traffic"], k["keyword_difficulty"],
               intent, k["vi"], k["date"]]
        for col, value in enumerate(row, start=1):
            ws.cell(row=i, column=col, value=value)

        change_cell = ws.cell(row=i, column=4)
        if k["is_new"] or (isinstance(k["change"], (int, float)) and k["change"] > 0):
            change_cell.font = green
        elif isinstance(k["change"], (int, float)) and k["change"] < 0:
            change_cell.font = red
        ws.cell(row=i, column=5).number_format = "#,##0"

    for col, width in enumerate([34, 12, 12, 9, 14, 10, 12, 26, 12, 13], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.auto_filter.ref = f"A{header_row}:J{ws.max_row}"

    buf = io.BytesIO()
    wb.save(buf)
    filename = f"keyword-report-{data['domain']}-{data['last_updated']}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------- pages

app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")


@app.get("/")
def index(request: Request):
    if not request.session.get("user"):
        return RedirectResponse("/login")
    return FileResponse(os.path.join(STATIC_DIR, "index.html"))


@app.get("/login")
def login_page(request: Request):
    if request.session.get("user"):
        return RedirectResponse("/")
    return FileResponse(os.path.join(STATIC_DIR, "login.html"))


def main():
    global DATA_FILE
    parser = argparse.ArgumentParser(description="Keyword ranking dashboard server")
    parser.add_argument("--data", help="Path to the results JSON file")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=8017)
    parser.add_argument(
        "--no-browser", action="store_true",
        help="Don't open the dashboard in a browser automatically",
    )
    args = parser.parse_args()

    if args.data:
        DATA_FILE = os.path.abspath(args.data)

    if DATA_FILE:
        source = DATA_FILE
    else:
        uri, db_name, coll_name = mongo_settings()
        source = f"MongoDB {uri} ({db_name}.{coll_name})"
        # Fail fast with a clear message if MongoDB isn't reachable.
        from pymongo import MongoClient

        try:
            MongoClient(uri, serverSelectionTimeoutMS=3000).server_info()
        except Exception as exc:
            raise SystemExit(
                f"Cannot connect to MongoDB at {uri} ({exc}).\n"
                "Is the MongoDB service running? Or serve a JSON file instead: "
                "python dashboard.py --data demo_results.json"
            )

    try:
        if get_db()["users"].count_documents({}) == 0:
            print("NOTE: no login accounts exist yet. Create one with:")
            print("  python manage_users.py add yourname --admin")
    except Exception:
        pass

    url = f"http://{args.host}:{args.port}"
    print(f"Serving data from: {source}")
    print(f"Dashboard: {url}  (Ctrl+C to stop)")

    if not args.no_browser:
        import threading
        import webbrowser

        threading.Timer(1.0, lambda: webbrowser.open(url)).start()

    import uvicorn

    uvicorn.run(app, host=args.host, port=args.port)


if __name__ == "__main__":
    main()
